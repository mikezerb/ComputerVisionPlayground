# --------------------------------------------------------- #
# Project 5. Exercise 2                                     #
#  Multiple object tracking using CSRT, KCF, MEDIANFLOW     #
#  Creating bounding box around moving objects              #
#  Calculating the coordinates of the                       #
#  centroid of the object                                   #
#  Evaluating and saving the results to excel file          #
# Michail Apostolidis                                       #
# --------------------------------------------------------- #

# import libraries
from __future__ import print_function
import sys
import cv2
from random import randint
from timeit import default_timer as timer
import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook


# Function for appending data frame (pandas) to excel file
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    if startrow is None:
        startrow = 0
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()


def CreateTrackerByName(trackerType):
    # Create a tracker based on tracker name
    if trackerType == trackerTypes[0]:
        tracker = cv2.TrackerCSRT_create()
    elif trackerType == trackerTypes[1]:
        tracker = cv2.TrackerKCF_create()
    elif trackerType == trackerTypes[2]:
        tracker = cv2.TrackerMedianFlow_create()
    else:
        tracker = None
        print('Incorrect tracker name')
        print('Available trackers are:')
        for t in trackerTypes:
            print(t)

    return tracker


# we use csrt, kcf and median tracking algorithm
trackerTypes = ['CSRT', 'KCF', 'MEDIANFLOW']

if __name__ == '__main__':
    for i in trackerTypes:
        print("\nTracking algorithm selected: " + i + "\n")
        trackerType = i

        # Set video to load
        videoPath = "./Multiple_object/basket.mp4"

        # Create a video capture object to read videos
        cap = cv2.VideoCapture(videoPath)

        # Read first frame
        success, frame = cap.read()
        # quit if unable to read the video file
        if not success:
            print('Failed to read video')
            sys.exit(1)

        ## Select boxes
        bboxes = []
        colors = []

        # OpenCV's selectROI function doesn't work for selecting multiple objects in Python
        # So we will call this function in a loop till we are done selecting all objects
        while True:
            # draw bounding boxes over objects
            # selectROI's default behaviour is to draw box starting from the center
            # when fromCenter is set to false, you can draw box starting from top left corner
            bbox = cv2.selectROI('MultiTracker', frame)
            bboxes.append(bbox)
            colors.append((randint(64, 255), randint(64, 255), randint(64, 255)))
            print("Press q to quit selecting boxes and start tracking")
            print("Press any other key to select next object")
            k = cv2.waitKey(0) & 0xFF
            if k == 113:  # q is pressed
                break

        print('Selected bounding boxes {}'.format(bboxes))

        # Initialize MultiTracker
        # No default algorithm specified
        tracker = cv2.MultiTracker()

        # Initialize MultiTracker with tracking algo
        # Specify tracker type

        # Create MultiTracker object
        multiTracker = cv2.MultiTracker_create()

        # Initialize MultiTracker
        for bbox in bboxes:
            multiTracker.add(CreateTrackerByName(trackerType), frame, bbox)

        frame_counter = 0  # counter of each frame
        # Calculating the centroid displacement:
        # init a 2D array -> [total_frames x number of objects]
        centroid_displacement = np.zeros((190, len(bboxes)))
        # define two 2D arrays for x, y of each object centroid
        center_x = np.zeros((190, len(bboxes)))
        center_y = np.zeros((190, len(bboxes)))
        # Process video and track objects
        # While loop for video
        while cap.isOpened():
            start_timer = timer()  # timer for performance of each frame
            # Reading the video stream
            success, frame = cap.read()
            if not success:
                break

            # get updated location of objects in subsequent frames
            success, boxes = multiTracker.update(frame)
            end_timer = timer()
            # draw tracked objects
            j = 0
            for i, newbox in enumerate(boxes):

                p1 = (int(newbox[0]), int(newbox[1]))
                p2 = (int(newbox[0] + newbox[2]), int(newbox[1] + newbox[3]))
                cv2.rectangle(frame, p1, p2, colors[i], 2, 1)

                # Finding the centroid
                # Draw circle in the center of the bounding box

                x2 = int(newbox[0]) + int(int(newbox[2]) / 2)
                y2 = int(newbox[1]) + int(int(newbox[3]) / 2)

                # adding the current coordinates to the corresponding array
                center_x[frame_counter][j] = x2
                center_y[frame_counter][j] = y2

                # draw the circle
                cv2.circle(frame, (x2, y2), 4, (0, 255, 0), -1)

                if frame_counter > 1:
                    print("Frame : ", frame_counter)
                    print("Previous center x : ", center_x[frame_counter - 1][j])
                    print("Current center x : ", center_x[frame_counter][j])
                    print("Previous center y : ", center_y[frame_counter - 1][j])
                    print("Current center y : ", center_y[frame_counter][j])
                    centroid_displacement[frame_counter][j] = (
                        abs(center_x[frame_counter - 1][j] - center_x[frame_counter][j] +
                            center_y[frame_counter - 1][j] - center_y[frame_counter][j]))
                elif frame_counter == 0:
                    centroid_displacement[frame_counter][j] = 0
                print("Centroid displacement : ", centroid_displacement[frame_counter][j])

                # display the centroid coordinates
                # on the center of the box of the frame
                text = "x: " + str(x2) + ", y: " + str(y2)
                cv2.putText(frame, text, (x2 - 10, y2 - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)
                # increment the loop counter
                j = j + 1

            # !Important Note :saving to file only works with two selected objects
            # Saving results to excel file
            # create pandas dataframe with the results
            results = pd.DataFrame([trackerType, frame_counter, end_timer - start_timer,
                                    centroid_displacement[frame_counter][0], centroid_displacement[frame_counter][1]])
            results = results.transpose()  # for saving to rows
            filename = 'multi-tracking_results.xlsx'  # name of the excel file
            # append dataframe to excel file for each evaluation
            append_df_to_excel(filename, results, header=None, index=False)

            # show frame
            cv2.imshow('MultiTracker', frame)
            frame_counter = frame_counter + 1
            # quit on ESC button
            if cv2.waitKey(1) & 0xFF == 27:  # Esc pressed
                break
        print("\n\nEnd of", trackerType, "for", len(bboxes), "objects...")
    # Close video stream
    cap.release()
    cv2.destroyAllWindows()
    # End of exercise 2
