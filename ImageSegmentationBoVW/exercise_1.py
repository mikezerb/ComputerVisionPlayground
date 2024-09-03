# --------------------------------------------------------- #
# Project 3. Exercise 1                                     #
#  Bag of Visual Words image classification                 #
#  Using Intel image classification data set                #
#  (images 150x150)                                         #
# Michail Apostolidis                                       #
# --------------------------------------------------------- #


# import necessary packages
import os
import cv2 as cv
import numpy as np
import time
import pandas as pd
from scipy.spatial import distance
from sklearn.cluster import MeanShift, estimate_bandwidth
from sklearn.cluster import MiniBatchKMeans  # KMeans
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.neighbors import KNeighborsClassifier  # train and evaluate the classifiers
from sklearn import preprocessing  # Convert Categorical Data For Scikit-Learn
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.naive_bayes import GaussianNB
from openpyxl import load_workbook


# Function for appending data frame (pandas) to excel file
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):

    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    if startrow is None:
        startrow = 0
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()


def load_data(folder):
    """
        Function for loading data (images) from dataset
        and returns the word dictionary that holds the images
        by category.

        :param folder: folder path of the dataset.
        :return: Dict.
    """
    images = {}
    size = 150
    sum = 0
    print("Loading... {}".format(folder))
    # Iterate through each folder corresponding to a category
    for filename in os.listdir(folder):
        count = 0
        category = []
        path = folder + filename
        print("In path = " + path)
        # Iterate through each image in our folder
        for cat in os.listdir(path):
            img = cv.imread(path + '/' + cat)  # reading the image
            if img is not None:
                img = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
                # Append the image to the output
                img = cv.resize(img, (size, size))  # resize the image
                category.append(img)
                count += 1
        sum += count
        images[filename] = category
        print("No of images parsed: ", count)
    print("Total No of images parsed: ", sum)
    return images


def extract_features(images, detect_type):
    """
        Function for creating descriptors using an detecting algorithm
        given in parameters and returns an array whose first index holds
        the decriptor_list without an order and the second index holds
        the detector_vectors dictionary which holds the descriptors but
        this is seperated class by class.

        :param images: word dictionary.
        :param detect_type: type of detection algorithm.
        :return: descriptor list and detector vectors.
    """
    print(' . start detecting points and calculating features for a given image set')
    detector_vectors = {}
    descriptor_list = []

    if detect_type == "sift":
        # Initiate SIFT detector
        detectorToUse = cv.xfeatures2d.SIFT_create()
    elif detect_type == "surf":
        # Initiate SURF detector
        detectorToUse = cv.xfeatures2d.SURF_create()
    elif detect_type == "orb":
        # Initiate ORB detector
        detectorToUse = cv.ORB_create()

    for nameOfCategory, availableImages in images.items():
        features = []
        tmpImgCounter = 1
        start = time.time()
        for img in availableImages:  # reminder: val
            kp, des = detectorToUse.detectAndCompute(img, None)
            if des is not None:  # set type as double for error
                des = des.astype(np.double)
                descriptor_list.extend(des)
                features.append(des)
            # print(".. image {:d} contributed:".format(tmpImgCounter), str(len(kp)), " points of interest")
            tmpImgCounter += 1
        print("Category {}, using {}".format(nameOfCategory, detect_type))
        print("Spend time:", time.time() - start)
        detector_vectors[nameOfCategory] = features

    print(' . finished detecting points and calculating features for a given image set')
    return [descriptor_list, detector_vectors]  # be aware of the []! this is ONE output as a list


def kmeansVisualWordsCreation(k, descriptor_list):
    """
        Function for using k-means clustering algorithm
        and returns an array which holds central points.

        :param k: number of cluster.
        :param descriptor_list: descriptors list(unordered 1d array).
        :return: central points array.
    """
    print(' . calculating central points for the existing feature values.')
    # kmeansModel = KMeans(n_clusters = k, n_init=10)
    batchSize = np.ceil(descriptor_list.__len__() / 50).astype('int')
    kmeansModel = MiniBatchKMeans(n_clusters=k, batch_size=batchSize, verbose=0).fit(descriptor_list)
    visualWords = kmeansModel.cluster_centers_  # a.k.a. centers of reference
    print(' . done calculating central points for the given feature set.')
    return visualWords, kmeansModel


# 3. building our own technic for creating histograms
# -- START --
def meanshiftVisualWordsCreation(k, descriptor_list):
    """
        Function for using Meanshift clustering algorithm
        and returns an array which holds central points.

        :param k: number of cluster.
        :param descriptor_list: descriptors list(unordered 1d array).
        :return: central points array.
    """
    print(' . calculating central points for the existing feature values.')
    print(" . using Meanshift clustering this will take a while...")
    # estimating bandwidth
    bandwidth = estimate_bandwidth(descriptor_list, quantile=0.2, n_samples=1000)
    # calculating meanshift model
    MeanShistModel = MeanShift(bandwidth=bandwidth, bin_seeding=True).fit(descriptor_list)
    labels = MeanShistModel.labels_
    labels_unique = np.unique(labels)
    n_clusters_ = len(labels_unique)
    print("n_clusters_ : ", n_clusters_)
    # getting visualwords from cluster centers
    visualWords = MeanShistModel.cluster_centers_
    print(' . done calculating central points for the given feature set.')
    return visualWords, MeanShistModel


def find_index(image, center):
    """
        Function for finding the closest center point of each descriptor
        with euclidean() method.

        :param image: feature descriptor.
        :param center: list of center points of the clustering algorithm.
        :return: index of the closest point.
    """
    count = 0
    index = 0
    for i in range(len(center)):
        if i == 0:
            count = distance.euclidean(image, center[i])
        else:
            dist = distance.euclidean(image, center[i])
            if dist < count:
                index = i
                count = dist
    return index


# 3. building our own technic for creating histograms
def calc_dict_histogram(dict, visualWords):
    """
        Function for creating a dictionary of the histograms for
        each image (separated by class).

        :param dict: dictionary of descriptors.
        :param visualWords: center points of the clustering algorithm.
        :return: histogram dictionary.
    """
    dictFeatures = {}
    for key, value in dict.items():
        category = []
        for img in value:
            Imghistogram = np.zeros(len(visualWords))
            for feat in img:
                i = find_index(feat, visualWords)   # getting index
                Imghistogram[i] += 1
            category.append(Imghistogram)  # append the histogram in each category
        dictFeatures[key] = category
    return dictFeatures

# -- END --


# Creation of the histograms. To create our each image by a histogram. We will create a vector of k values for each
# image. For each keypoints in an image, we will find the nearest center, defined using training set
# and increase by one its value
def mapFeatureValsToHistogram(DataFeaturesByClass, visualWords, TrainedKmeansModel):
    # depending on the approach you may not need to use all inputs
    histogramsList = []
    targetClassList = []
    numberOfBinsPerHistogram = visualWords.shape[0]

    for categoryIdx, featureValues in DataFeaturesByClass.items():
        for tmpImageFeatures in featureValues:  # yes, we check one by one the values in each image for all images
            tmpImageHistogram = np.zeros(numberOfBinsPerHistogram)
            tmpIdx = list(TrainedKmeansModel.predict(tmpImageFeatures))
            clusterValue, visualWordMatchCounts = np.unique(tmpIdx, return_counts=True)
            tmpImageHistogram[clusterValue] = visualWordMatchCounts
            # do not forget to normalize the histogram values
            numberOfDetectedPointsInThisImage = tmpIdx.__len__()
            tmpImageHistogram = tmpImageHistogram / numberOfDetectedPointsInThisImage

            # now update the input and output coresponding lists
            histogramsList.append(tmpImageHistogram)
            targetClassList.append(categoryIdx)

    return histogramsList, targetClassList


def evaluate(y_train, y_pred_train, y_test, y_pred_test, feature_type, cluster_type, ratio, classifier):
    """
        Function for evaluating the bag of visual words algorithm
        and saving the results to an excel file.
    """
    # calculating the scores (accuracy, precision, recall, f1)
    acc_train = accuracy_score(y_train, y_pred_train)
    acc_test = accuracy_score(y_test, y_pred_test)
    pre_train = precision_score(y_train, y_pred_train, average='macro')
    pre_test = precision_score(y_test, y_pred_test, average='macro')
    rec_train = recall_score(y_train, y_pred_train, average='macro')
    rec_test = recall_score(y_test, y_pred_test, average='macro')
    f1_train = f1_score(y_train, y_pred_train, average='macro')
    f1_test = f1_score(y_test, y_pred_test, average='macro')

    # print the scores to console
    print('Accuracy scores of K-NN classifier are:',
          'train: {:.2f}'.format(acc_train), 'and test: {:.2f}.'.format(acc_test))
    print('Precision scores of Logistic regression classifier are:',
          'train: {:.2f}'.format(pre_train), 'and test: {:.2f}.'.format(pre_test))
    print('Recall scores of K-NN classifier are:',
          'train: {:.2f}'.format(rec_train), 'and test: {:.2f}.'.format(rec_test))
    print('F1 scores of K-NN classifier are:',
          'train: {:.2f}'.format(f1_train), 'and test: {:.2f}.'.format(f1_test))
    print('')

    # Saving results to excel file
    # create pandas dataframe with the results
    results = pd.DataFrame([feature_type, cluster_type, ratio, classifier,
                            acc_train, pre_train, rec_train, f1_train,
                            acc_test, pre_test, rec_test, f1_test])

    results = results.transpose()   # for saving to rows

    filename = 'results.xlsx'    # name of the excel file

    # append dataframe to excel file for each evaluation
    append_df_to_excel(filename, results, header=None, index=False)


if __name__ == '__main__':
    # 1. Load dataset with train and test images.
    # paths for train and test data set images
    # starting with 80% train 20% test and then run it with 60% and 40%
    TrainImagePath = './80-20/train/'
    TestImagePath = './80-20/test/'

    # loading the train images
    trainImages = load_data(TrainImagePath)

    # 2. Extract features using sift and 2 other algorithms of our choosing (ORB, SURF)
    # calculate points and descriptor values per image
    trainDataFeatures = extract_features(trainImages, "sift")
    # Takes the descriptor list which is unordered one
    TrainDescriptorList = trainDataFeatures[0]

    # A. K-means clustering:
    # create the central points for the histograms using k means.
    # here we use a rule of the thumb to create the expected number of cluster centers
    numberOfClasses = trainImages.__len__()  # retrieve num of classes from dictionary
    possibleNumOfCentersToUse = 10 * numberOfClasses
    visualWords, TrainedKmeansModel = kmeansVisualWordsCreation(possibleNumOfCentersToUse, TrainDescriptorList)

    # Takes the sift feature values that is seperated class by class for train data, we need this to calculate the
    # histograms
    trainBoVWFeatureVals = trainDataFeatures[1]

    # B. Mean Shift
    # 3. Other method for calculating histograms:
    # creating histogram for the train data
    BoVWTrain = calc_dict_histogram(trainBoVWFeatureVals, visualWords)
    # Takes the sift features that is seperated class by class for test data
    testBovWFeatures = extract_features(load_data(TestImagePath), "sift")[1]
    # Creates histograms for test data
    bovw_test = calc_dict_histogram(testBovWFeatures, visualWords)

    # create the train input train output format
    trainHistogramsList, trainTargetsList = mapFeatureValsToHistogram(trainBoVWFeatureVals, visualWords,
                                                                      TrainedKmeansModel)

    X_train = np.stack(trainHistogramsList, axis=0)

    # Create a label (category) encoder object
    labelEncoder = preprocessing.LabelEncoder()
    labelEncoder.fit(trainTargetsList)
    # convert the categories from strings to names
    y_train = labelEncoder.transform(trainTargetsList)

    # Train the classifiers and evaluate the accuracy:
    knn = KNeighborsClassifier()
    knn.fit(X_train, y_train)
    print('Accuracy of K-NN classifier on training set: {:.2f}'.format(knn.score(X_train, y_train)))

    clf = DecisionTreeClassifier().fit(X_train, y_train)
    print('Accuracy of Decision Tree classifier on training set: {:.2f}'.format(clf.score(X_train, y_train)))

    gnb = GaussianNB()
    gnb.fit(X_train, y_train)
    print('Accuracy of GNB classifier on training set: {:.2f}'.format(gnb.score(X_train, y_train)))

    svm = SVC()  # make classifier object
    svm.fit(X_train, y_train)  # train the model
    print('Accuracy of SVM classifier on training set: {:.2f}'.format(svm.score(X_train, y_train)))

    # -------------------------------------------------------------
    # Using the visual words to the test data:

    del trainImages, trainBoVWFeatureVals, trainDataFeatures, TrainDescriptorList  # clear some space

    # loading the test images
    testImages = load_data(TestImagePath)  # take all images category by category for train set

    # calculating points and descriptor values per image
    testDataFeatures = extract_features(testImages, "sift")

    # Takes each feature values that is seperated class by class for train data, to calculate the histograms
    testBoVWFeatureVals = testDataFeatures[1]

    # creating the test input / test output format
    testHistogramsList, testTargetsList = mapFeatureValsToHistogram(testBoVWFeatureVals, visualWords,
                                                                    TrainedKmeansModel)
    X_test = np.array(testHistogramsList)
    y_test = labelEncoder.transform(testTargetsList)

    # 4. Evaluating each feature extraction and clustering detection
    # print the scores
    print('\nPrinting performance scores:\n')

    # Classification tree
    # predict outcomes for test data and calculate the test scores
    y_pred_train = clf.predict(X_train)
    y_pred_test = clf.predict(X_test)
    evaluate(y_train, y_pred_train, y_test, y_pred_test, "SIFT", "Meanshift", "80/20", "Decision Tree")

    # KNN predictions
    # now check for both train and test data, how well the model learned the patterns
    y_pred_train = knn.predict(X_train)
    y_pred_test = knn.predict(X_test)
    evaluate(y_train, y_pred_train, y_test, y_pred_test, "SIFT", "Meanshift", "80/20", "KNN predictions")

    # naive Bayes
    # now check for both train and test data, how well the model learned the patterns
    y_pred_train = gnb.predict(X_train)
    y_pred_test = gnb.predict(X_test)
    evaluate(y_train, y_pred_train, y_test, y_pred_test, "SIFT", "Meanshift", "80/20", "naive Bayes")

    # support vector machines
    # now check for both train and test data, how well the model learned the patterns
    y_pred_train = gnb.predict(X_train)
    y_pred_test = gnb.predict(X_test)
    evaluate(y_train, y_pred_train, y_test, y_pred_test, "SIFT", "Meanshift", "80/20", "SVM classifier")
