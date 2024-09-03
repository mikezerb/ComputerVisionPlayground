# --------------------------------------------------------- #
# Project 4. Exercise 1                                     #
#  CNN image classification                                 #
#  Using Intel image classification data set                #
#  (images 150x150)                                         #
# Michail Apostolidis                                       #
# --------------------------------------------------------- #

# import necessary packages
import keras
from keras import backend as K
from keras.datasets import mnist
from keras.models import Sequential
from keras.layers import Dense, Dropout, Flatten
from keras.layers import Conv2D, MaxPooling2D
import matplotlib.pyplot as plt
import random
from sklearn.svm import SVC
import numpy as np
from sklearn.metrics import classification_report
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import pandas as pd
from openpyxl import load_workbook
import os


# Function for appending data frame (pandas) to excel file
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    if startrow is None:
        startrow = 0
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()


def evaluate(y_train, y_pred_train, y_test, y_pred_test, feature_type, cluster_type, ratio, classifier):
    """
        Function for evaluating CNN  nd saving the results to an excel file.
    """
    # calculating the scores (accuracy, precision, recall, f1)
    acc_train = accuracy_score(y_train, y_pred_train)
    acc_test = accuracy_score(y_test, y_pred_test)
    pre_train = precision_score(y_train, y_pred_train, average='macro')
    pre_test = precision_score(y_test, y_pred_test, average='macro')
    rec_train = recall_score(y_train, y_pred_train, average='macro')
    rec_test = recall_score(y_test, y_pred_test, average='macro')
    f1_train = f1_score(y_train, y_pred_train, average='macro')
    f1_test = f1_score(y_test, y_pred_test, average='macro')

    # Saving results to excel file
    # create pandas dataframe with the results
    results = pd.DataFrame([feature_type, cluster_type, ratio, classifier,
                            acc_train, pre_train, rec_train, f1_train,
                            acc_test, pre_test, rec_test, f1_test])

    results = results.transpose()  # for saving to rows

    filename = 'results.xlsx'  # name of the excel file

    # append dataframe to excel file for each evaluation
    append_df_to_excel(filename, results, header=None, index=False)


def recall_m(y_true, y_pred):
    true_positives = K.sum(K.round(K.clip(y_true * y_pred, 0, 1)))
    possible_positives = K.sum(K.round(K.clip(y_true, 0, 1)))
    recall = true_positives / (possible_positives + K.epsilon())
    return recall


def precision_m(y_true, y_pred):
    true_positives = K.sum(K.round(K.clip(y_true * y_pred, 0, 1)))
    predicted_positives = K.sum(K.round(K.clip(y_pred, 0, 1)))
    precision = true_positives / (predicted_positives + K.epsilon())
    return precision


def f1_m(y_true, y_pred):
    precision = precision_m(y_true, y_pred)
    recall = recall_m(y_true, y_pred)
    return 2 * ((precision * recall) / (precision + recall + K.epsilon()))


if __name__ == '__main__':
    # get the data
    from DataLoadClassif import X_train, Y_train, X_test, Y_test, X_val, Y_val, IMG_HEIGHT, IMG_WIDTH, IMG_CHANNELS

    batch_size = 100
    num_classes = np.unique(Y_train).__len__()
    epochs = 15

    # define some CNN parameters
    baseNumOfFilters = 16

    # the data, split between train and test sets
    # (X_train, Y_train), (x_test, y_test) = mnist.load_data()

    print('X_train shape:', X_train.shape)
    print(X_train.shape[0], 'train samples')
    print(X_test.shape[0], 'test samples')

    # convert class vectors to binary class matrices

    Y_train = keras.utils.to_categorical(Y_train, num_classes)
    Y_test = keras.utils.to_categorical(Y_test, num_classes)
    Y_val = keras.utils.to_categorical(Y_val, num_classes)

    # here we define and load the model

    inputs = keras.layers.Input((IMG_HEIGHT, IMG_WIDTH, IMG_CHANNELS))
    s = keras.layers.Lambda(lambda x: x / 255)(inputs)  # normalize the input
    conv1 = keras.layers.Conv2D(filters=baseNumOfFilters, kernel_size=(13, 13))(s)
    pool1 = keras.layers.MaxPooling2D(pool_size=(2, 2))(conv1)
    conv2 = keras.layers.Conv2D(filters=baseNumOfFilters * 2, kernel_size=(7, 7))(pool1)
    pool2 = keras.layers.MaxPooling2D(pool_size=(2, 2))(conv2)
    conv3 = keras.layers.Conv2D(filters=baseNumOfFilters * 4, kernel_size=(3, 3))(pool2)
    drop3 = keras.layers.Dropout(0.25)(conv3)
    flat1 = keras.layers.Flatten()(drop3)
    dense1 = keras.layers.Dense(128, activation='relu')(flat1)
    outputs = keras.layers.Dense(Y_train.shape[1], activation='softmax')(dense1)

    CNNmodel = keras.Model(inputs=[inputs], outputs=[outputs])
    CNNmodel.compile(optimizer='sgd', loss=keras.losses.categorical_crossentropy, metrics=['accuracy'])
    # print model summary
    CNNmodel.summary()

    # fit model parameters, given a set of training data
    callbacksOptions = [
        keras.callbacks.EarlyStopping(patience=15, verbose=1),
        keras.callbacks.ReduceLROnPlateau(factor=0.1, patience=5, min_lr=0.0001, verbose=1),
        keras.callbacks.ModelCheckpoint('tmpCNN.h5', verbose=1, save_best_only=True, save_weights_only=True)]

    CNNmodel.fit(X_train, Y_train, batch_size=batch_size, shuffle=True, epochs=epochs, verbose=1,
                 callbacks=callbacksOptions, validation_data=(X_val, Y_val))

    # calculate some common performance scores
    score = CNNmodel.evaluate(X_test, Y_test, verbose=0)
    print('Test loss:', score[0])
    print('Test accuracy:', score[1])

    y_pred = CNNmodel.predict(X_test, batch_size=64, verbose=1)
    y_pred_bool = np.argmax(y_pred, axis=1)

    # saving the trained model
    model_name = 'intelCNN.h5'
    CNNmodel.save(model_name)

    # # loading a trained model & use it over test data
    loaded_model = keras.models.load_model(model_name)

    y_test_predictions_vectorized = loaded_model.predict(X_test)
    y_test_predictions = np.argmax(y_test_predictions_vectorized, axis=1)

    # Print f1, precision, and recall scores
    print(precision_score(y_test_predictions_vectorized, y_test_predictions, average="macro"))
    print(recall_score(y_test_predictions_vectorized, y_test_predictions, average="macro"))
    print(f1_score(y_test_predictions_vectorized, y_test_predictions, average="macro"))

    print(classification_report(y_pred_bool, y_test_predictions))
    print(classification_report(y_test_predictions, y_pred_bool))
    print(recall_m(y_pred_bool, y_test_predictions))
